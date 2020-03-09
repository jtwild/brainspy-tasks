import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook


class ExcelFile():
    def __init__(self, file_path, overwrite = True):
        self.overwrite = overwrite
        if os.path.exists(file_path) and os.path.isfile(file_path):
            if self.overwrite:
                os.remove(file_path)
            else:
                # Check if we can write to the file. If we cant,
                try:
                    with open(file_path, 'a') as a:
                        pass
                except IOError:
                    raise IOError('Excel file cannot be opened and written to! Is it in use?')
                #if not os.access(file_path, os.W_OK):  # this does not work because only checks permissions.
        else:
            # If the file does not exist, we need to call ExcelWriter with the 'w' function, a.k.a. 'overwriting' a nonexistent file.
            self.overwrite = True
        self.file_path = file_path
        self.writer = None
        # self.writer = pd.ExcelWriter(file_path, engine='openpyxl')  # pylint: disable=abstract-class-instantiated

    def read_file(self, sheet_name='main'):
        xlsx = pd.ExcelFile(self.file_path, engine='openpyxl')
        return xlsx.parse(sheet_name)

    def init_data(self, column_names, index=None):
        if index is None:
            self.data = pd.DataFrame(columns=column_names)
        else:
            self.data = pd.DataFrame(index=pd.Series(map(str, index)), columns=column_names)

    def insert_column(self, column_name, column_data):
        self.data[column_name] = pd.Series(column_data, index=self.data.index)

    def save_tab(self, tab_name, data=None):
        if data is None:
            aux = self.data
        else:
            aux = data
        if not self.overwrite:
            # Append at end of existing sheet
            row_index = self.writer.book[tab_name].max_row
            header_mode = None
        else:
            # Overwrite, default behaviour.
            row_index = 0
            header_mode = True
        aux.to_excel(self.writer, sheet_name=tab_name, header=header_mode, startrow=row_index)
        self.writer.save()

    def reset(self):
        if self.writer is not None:
            self.close_file()
        self.open_file()

    def open_file(self):
        book = None
        if os.path.exists(self.file_path) and os.path.isfile(self.file_path):
            book = load_workbook(self.file_path)
        # Check if we should write ('w') or append ('w') data:
        if self.overwrite:
            overwrite_mode = 'w'
        else:
            overwrite_mode = 'a'
        # Engine changed from xlsxwriter to openpyxl to fix no attribute errors.
        self.writer = pd.ExcelWriter(self.file_path, engine='openpyxl', mode=overwrite_mode)  # pylint: disable=abstract-class-instantiated
        if book is not None:
            self.writer.book = book
            self.writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    def close_file(self):
        self.writer.save()
        self.writer.close()

    def add_result(self, results, label=None):
        if label is None:
            self.data = self.data.append(pd.Series(results), ignore_index=True)
        else:
            self.data.loc[str(label)] = pd.Series(results)


def get_series_with_numpy(series):
    return series.apply(lambda x:
                        np.fromstring(
                            x.replace('\n', '')
                            .replace('[', '')
                            .replace(']', '')
                            .replace('  ', ' '), sep=' '))


def load_bn_values(excel):
    bn_statistics = {'bn_1': {}, 'bn_2': {}}
    bn_statistics['bn_1']['mean'] = excel['bn_1_mean']
    bn_statistics['bn_1']['var'] = excel['bn_1_var']
    bn_statistics['bn_2']['mean'] = excel['bn_2_mean']
    bn_statistics['bn_2']['var'] = excel['bn_2_var']
    return bn_statistics
